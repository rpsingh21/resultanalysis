import os

from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect, Http404

from openpyxl import Workbook

from result.models import (
	College,
	Branch,
	Course,
	TotalMarks,
	Subject,
	Student,
	Marks,
	)

# Create your views here.

def semseterResultxlsx(request,collegeCode,branchCode,yearOfJoining,semester):
	if not request.user.is_authenticated:
		raise Http404
		
	# GET DATA OF CALSS STUDENT 
	totalMarks = TotalMarks.objects.filter(student__college__collegeCode=collegeCode, student__branchCode__branchCode=branchCode, student__yearOfJoining = yearOfJoining, semester=semester).order_by('-totalMarks')

	# GET SUBJECT LIST 
	subjectList=['ROLL NO.','STUDENT NAME']
	try:
		marks=Marks.objects.filter(student=totalMarks[0].student,subjectCode__semester=semester).order_by('subjectCode')
	except Exception as e:
		raise Http404
	next=len(marks)
	for _ in range(2):
		for i in marks:
			subjectList.append(i.subjectCode.subjectCode)
	subjectList += ["Total Internal","Total External","Total"]
	fileName = file_path = os.path.join(settings.MEDIA_ROOT, collegeCode+"."+branchCode+"."+yearOfJoining+"."+semester+".xlsx")
	
	# WRITE WORKBOOK 
	workBook = Workbook()
	workSheet= workBook.active
	workSheet.title="Semester "+semester+" Result"
	workSheet.append(subjectList)
	subjectList=subjectList[2:next+2]
	row=2
	for totalMark in totalMarks:
		col=3
		marks=Marks.objects.filter(student=totalMark.student,subjectCode__subjectCode__in=subjectList).order_by('subjectCode')
		if marks.exists():
			workSheet.cell(row=row, column=1, value=totalMark.student.rollNo)
			workSheet.cell(row=row, column=2, value=totalMark.student.name)
			for mark in marks:
				if mark.internal != -1:
					workSheet.cell(row=row, column=col, value=mark.internal)
				else:
					workSheet.cell(row=row, column=col, value="--")
				if mark.external != -1:
					workSheet.cell(row=row, column=col+next, value=mark.external)
				else:
					workSheet.cell(row=row, column=col+next, value="--")
				col += 1
			workSheet.cell(row=row, column=col+next, value=totalMark.internal)
			workSheet.cell(row=row, column=col+next+1, value=totalMark.external)
			workSheet.cell(row=row, column=col+next+2, value=totalMark.totalMarks)
			row += 1
			
	f=workBook.save(filename=fileName)
	if os.path.exists(fileName):
		with open(fileName, 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
			response['Content-Disposition'] = 'inline; filename=' + os.path.basename(fileName)
			return response
	raise Http404


